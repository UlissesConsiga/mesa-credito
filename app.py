import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
import json
from datetime import datetime
from zoneinfo import ZoneInfo
import pandas as pd
import matplotlib.pyplot as plt
import time
import bcrypt
from functools import wraps
import logging

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

st.set_page_config(
    layout="wide",
    initial_sidebar_state="expanded",
    page_title="Consiga Empréstimos",
    page_icon="💼"
)

# ==============================
# CONFIGURAÇÕES OTIMIZADAS
# ==============================

# Cache mais longo para reduzir requisições
CACHE_TTL_LONGO = 1800  # 30 minutos
CACHE_TTL_MEDIO = 600   # 10 minutos
CACHE_TTL_CURTO = 180   # 3 minutos

# Controle de rate limiting
MAX_REQUESTS_PER_MINUTE = 30
request_times = []

# Tema removido - usando design profissional fixo

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ============================= */
/* VARIÁVEIS */
/* ============================= */
:root {
    --green:        #16a34a;
    --green-light:  #22c55e;
    --green-dark:   #15803d;
    --orange:       #f97316;
    --orange-dark:  #ea580c;
    --sidebar-bg:   #1a1f2e;
}

/* ============================= */
/* BASE */
/* ============================= */
* { font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important; }

*, *:hover { cursor: default !important; }
a, button, [role="button"], .stButton>button,
select, [data-testid="stSelectbox"], [data-baseweb="select"], label[for]
    { cursor: pointer !important; }
input[type="text"], input[type="password"], textarea { cursor: text !important; }

header[data-testid="stHeader"] { display: none !important; }
[data-testid="stToolbar"]       { display: none !important; }
#MainMenu                       { display: none !important; }
footer                          { display: none !important; }

/* ============================= */
/* MODO CLARO */
/* ============================= */
@media (prefers-color-scheme: light) {
    .stApp, [data-testid="stAppViewContainer"] { background: #f8fafc !important; }
    h1, h2, h3 { color: #15803d !important; }
    p, li { color: #1e293b !important; }
    label { color: #1e293b !important; }
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] span { color: #1e293b !important; }
    .stTextInput input, .stNumberInput input, .stTextArea textarea {
        background: #ffffff !important;
        border: 1.5px solid #cbd5e1 !important;
        color: #0f172a !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
        border-color: #16a34a !important;
        box-shadow: 0 0 0 3px rgba(22,163,74,0.12) !important;
    }
    .stTextInput input::placeholder, .stTextArea textarea::placeholder { color: #94a3b8 !important; }
    .stSelectbox div[data-baseweb="select"] {
        background: #ffffff !important;
        border: 1.5px solid #cbd5e1 !important;
    }
    .stSelectbox div[data-baseweb="select"] * { color: #0f172a !important; }
    .stDateInput input { background: #ffffff !important; border: 1.5px solid #cbd5e1 !important; color: #0f172a !important; }
    [data-testid="stDataFrame"] table { background: #ffffff !important; }
    [data-testid="stDataFrame"] tbody tr:nth-child(even) { background: #f1f5f9 !important; }
    [data-testid="stDataFrame"] tbody tr:hover { background: #dcfce7 !important; }
    [data-testid="stDataFrame"] tbody td { color: #1e293b !important; }
    [data-testid="stMetricValue"] { color: #15803d !important; }
    [data-testid="stMetricLabel"] { color: #475569 !important; }
    .stRadio label { color: #1e293b !important; }
}

/* ============================= */
/* MODO ESCURO */
/* ============================= */
@media (prefers-color-scheme: dark) {
    .stApp, [data-testid="stAppViewContainer"] {
        background: linear-gradient(160deg, #0f172a 0%, #1e293b 100%) !important;
    }
    h1, h2, h3 { color: #22c55e !important; }
    p, li { color: #f1f5f9 !important; }
    label { color: #e2e8f0 !important; }
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] span { color: #f1f5f9 !important; }
    .stTextInput input, .stNumberInput input, .stTextArea textarea {
        background: #1e293b !important;
        border: 1.5px solid #334155 !important;
        color: #f1f5f9 !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {
        border-color: #22c55e !important;
        box-shadow: 0 0 0 3px rgba(34,197,94,0.2) !important;
    }
    .stSelectbox div[data-baseweb="select"] {
        background: #1e293b !important;
        border: 1.5px solid #334155 !important;
    }
    .stSelectbox div[data-baseweb="select"] * { color: #f1f5f9 !important; }
    [data-testid="stDataFrame"] table { background: #1e293b !important; }
    [data-testid="stDataFrame"] tbody tr:nth-child(even) { background: #0f172a !important; }
    [data-testid="stDataFrame"] tbody tr:hover { background: #14532d !important; }
    [data-testid="stDataFrame"] tbody td { color: #f1f5f9 !important; }
    [data-testid="stMetricValue"] { color: #22c55e !important; }
    [data-testid="stMetricLabel"] { color: #94a3b8 !important; }
}

/* ============================= */
/* TÍTULOS */
/* ============================= */
h1 { font-weight: 800 !important; font-size: 2.25rem !important; margin-bottom: 1.5rem !important; letter-spacing: -0.025em !important; }
h2 { font-weight: 700 !important; font-size: 1.75rem !important; margin-top: 1.5rem !important; margin-bottom: 0.75rem !important; }
h3 { font-weight: 600 !important; font-size: 1.375rem !important; }

/* ============================= */
/* LABELS */
/* ============================= */
label { font-weight: 600 !important; font-size: 0.8125rem !important; text-transform: uppercase !important; letter-spacing: 0.06em !important; }

/* ============================= */
/* INPUTS */
/* ============================= */
.stTextInput input, .stNumberInput input, .stTextArea textarea {
    border-radius: 10px !important;
    padding: 12px 16px !important;
    font-size: 0.9375rem !important;
    font-weight: 500 !important;
    transition: border-color 0.15s, box-shadow 0.15s !important;
}
.stSelectbox div[data-baseweb="select"] { border-radius: 10px !important; }

/* ============================= */
/* BOTÕES */
/* ============================= */
.stButton>button {
    background: linear-gradient(135deg, #22c55e 0%, #16a34a 100%) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 12px 24px !important;
    font-size: 0.9375rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.02em !important;
    transition: all 0.15s ease !important;
    box-shadow: 0 2px 8px rgba(22,163,74,0.25) !important;
}
.stButton>button:hover {
    background: linear-gradient(135deg, #16a34a 0%, #15803d 100%) !important;
    box-shadow: 0 4px 16px rgba(22,163,74,0.35) !important;
    transform: translateY(-1px) !important;
}
.stButton>button:active { transform: translateY(0) !important; }
.stButton>button[kind="primary"] {
    background: linear-gradient(135deg, #f97316 0%, #ea580c 100%) !important;
    box-shadow: 0 2px 8px rgba(249,115,22,0.25) !important;
}
.stButton>button[kind="primary"]:hover {
    background: linear-gradient(135deg, #ea580c 0%, #c2410c 100%) !important;
}
.stDownloadButton>button {
    background: linear-gradient(135deg, #22c55e 0%, #16a34a 100%) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
}

/* ============================= */
/* SIDEBAR — fixa, sem botão de colapsar */
/* ============================= */
[data-testid="stSidebar"] {
    border-right: 1px solid rgba(0,0,0,0.06) !important;
    box-shadow: 4px 0 16px rgba(0,0,0,0.08) !important;
    min-width: 240px !important;
}
[data-testid="stSidebar"] img {
    border-radius: 6px !important;
    background: transparent !important;
    padding: 0 !important;
    display: block !important;
    margin: 0 auto !important;
}

/* Sidebar — MODO CLARO: verde bem clarinho */
@media (prefers-color-scheme: light) {
    [data-testid="stSidebar"] {
        background: #f0fdf4 !important;
        border-right: 1px solid #bbf7d0 !important;
    }
    [data-testid="stSidebar"] * { color: #14532d !important; }
    [data-testid="stSidebar"] label {
        color: #166534 !important;
        font-size: 0.7rem !important;
        letter-spacing: 0.08em !important;
    }
    [data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] {
        background: #ffffff !important;
        border: 1px solid #86efac !important;
        border-radius: 8px !important;
    }
    [data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] * {
        color: #14532d !important;
    }
    [data-testid="stSidebar"] .stButton>button {
        background: #16a34a !important;
        border: none !important;
        color: #ffffff !important;
        box-shadow: 0 2px 6px rgba(22,163,74,0.25) !important;
    }
    [data-testid="stSidebar"] .stButton>button:hover {
        background: #15803d !important;
        transform: none !important;
    }
    [data-testid="stSidebar"] hr {
        background: #bbf7d0 !important;
    }
}

/* Sidebar — MODO ESCURO: azul-escuro */
@media (prefers-color-scheme: dark) {
    [data-testid="stSidebar"] {
        background: #1a1f2e !important;
        border-right: 1px solid rgba(255,255,255,0.05) !important;
        box-shadow: 4px 0 20px rgba(0,0,0,0.25) !important;
    }
    [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] label {
        color: #64748b !important;
        font-size: 0.7rem !important;
        letter-spacing: 0.08em !important;
    }
    [data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] {
        background: rgba(255,255,255,0.05) !important;
        border: 1px solid rgba(255,255,255,0.1) !important;
        border-radius: 8px !important;
    }
    [data-testid="stSidebar"] .stButton>button {
        background: rgba(34,197,94,0.12) !important;
        border: 1px solid rgba(34,197,94,0.25) !important;
        color: #22c55e !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] .stButton>button:hover {
        background: rgba(34,197,94,0.22) !important;
        border-color: rgba(34,197,94,0.45) !important;
        transform: none !important;
    }
}

/* Esconde o botão de colapsar — todos os seletores possíveis */
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapseButton"],
button[data-testid="baseButton-header"],
section[data-testid="stSidebar"] > div > button {
    display: none !important;
    visibility: hidden !important;
    pointer-events: none !important;
    width: 0 !important;
    min-width: 0 !important;
    height: 0 !important;
    overflow: hidden !important;
    position: absolute !important;
    left: -9999px !important;
}

/* ============================= */
/* MÉTRICAS, DATAFRAMES, MISC */
/* ============================= */
[data-testid="stMetricValue"] { font-size: 2rem !important; font-weight: 800 !important; }
[data-testid="stMetricLabel"] { font-weight: 600 !important; text-transform: uppercase !important; letter-spacing: 0.05em !important; font-size: 0.7rem !important; }

[data-testid="stDataFrame"] { border-radius: 12px !important; overflow: hidden !important; box-shadow: 0 4px 16px rgba(0,0,0,0.08) !important; }
[data-testid="stDataFrame"] thead tr th {
    background: linear-gradient(135deg, #22c55e 0%, #16a34a 100%) !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    padding: 14px 16px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
    font-size: 0.7rem !important;
    border: none !important;
}
[data-testid="stDataFrame"] tbody td { padding: 12px 16px !important; font-weight: 500 !important; }

hr { border: none !important; height: 1px !important; background: linear-gradient(90deg, transparent, rgba(100,116,139,0.3), transparent) !important; margin: 2rem 0 !important; }
.stRadio > label { font-weight: 600 !important; }
.stSpinner > div { border-top-color: #22c55e !important; }

</style>
""", unsafe_allow_html=True)

# Header removido - design profissional aplicado

# JavaScript para fixar a sidebar permanentemente (impede colapso)
st.markdown("""
<script>
(function fixSidebar() {
    function forceExpand() {
        // Remove o atributo aria-expanded=false que colapsa a sidebar
        var sidebar = window.parent.document.querySelector('[data-testid="stSidebar"]');
        if (sidebar) {
            sidebar.removeAttribute('aria-hidden');
            sidebar.style.display = '';
            sidebar.style.visibility = 'visible';
            sidebar.style.width = '';
            sidebar.style.minWidth = '';
        }

        // Esconde e desabilita o botão de colapsar
        var btns = window.parent.document.querySelectorAll(
            '[data-testid="collapsedControl"], [data-testid="stSidebarCollapseButton"]'
        );
        btns.forEach(function(btn) {
            btn.style.display = 'none';
            btn.style.pointerEvents = 'none';
            // Bloqueia cliques
            btn.addEventListener('click', function(e) {
                e.stopImmediatePropagation();
                e.preventDefault();
            }, true);
        });
    }

    // Executa imediatamente e observa mudanças no DOM
    forceExpand();
    var observer = new MutationObserver(forceExpand);
    observer.observe(window.parent.document.body, { childList: true, subtree: true, attributes: true });
})();
</script>
""", unsafe_allow_html=True)

# ==============================
# FUNÇÕES DE UTILIDADE
# ==============================

def rate_limit_check():
    """Controla o rate limiting de requisições"""
    global request_times
    now = time.time()
    # Remove requisições antigas (mais de 1 minuto)
    request_times = [t for t in request_times if now - t < 60]
    
    if len(request_times) >= MAX_REQUESTS_PER_MINUTE:
        wait_time = 60 - (now - request_times[0])
        if wait_time > 0:
            logger.warning(f"Rate limit atingido. Aguardando {wait_time:.1f}s")
            time.sleep(wait_time)
            request_times = []
    
    request_times.append(now)

def retry_on_failure(max_retries=3, delay=2):
    """Decorator para retry automático em caso de falha"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    rate_limit_check()
                    return func(*args, **kwargs)
                except gspread.exceptions.APIError as e:
                    logger.error(f"Erro API Google Sheets (tentativa {attempt + 1}/{max_retries}): {e}")
                    if attempt < max_retries - 1:
                        time.sleep(delay * (attempt + 1))  # Backoff exponencial
                    else:
                        st.error("⚠ Erro ao conectar com Google Sheets. Tente novamente em alguns segundos.")
                        raise
                except Exception as e:
                    logger.error(f"Erro inesperado: {e}")
                    if attempt < max_retries - 1:
                        time.sleep(delay)
                    else:
                        raise
            return None
        return wrapper
    return decorator

def hash_senha(senha):
    """Criptografa senha usando bcrypt"""
    return bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verificar_senha(senha, hash_armazenado):
    """Verifica se a senha corresponde ao hash"""
    try:
        return bcrypt.checkpw(senha.encode('utf-8'), hash_armazenado.encode('utf-8'))
    except:
        # Fallback para senhas antigas em texto plano
        return senha == hash_armazenado

# ==============================
# CONEXÃO COM GOOGLE SHEETS
# ==============================

SHEET_NAME = os.environ.get("SHEET_NAME", "")
google_creds_str = os.environ.get("GOOGLE_CREDENTIALS", "{}")

try:
    google_creds = json.loads(google_creds_str)
except json.JSONDecodeError:
    st.error("⚠ Erro ao carregar credenciais do Google. Verifique a variável GOOGLE_CREDENTIALS.")
    st.stop()

scope = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]

@st.cache_resource(ttl=3600)  # Cache de 1 hora para conexão
def conectar_google():
    """Conecta ao Google Sheets com tratamento de erros"""
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_dict(google_creds, scope)
        client = gspread.authorize(creds)
        planilha = client.open(SHEET_NAME)
        sheet = planilha.worksheet("BASE_CONTROLE")
        sheet_usuarios = planilha.worksheet("USUARIOS")
        logger.info("Conexão com Google Sheets estabelecida com sucesso")
        return sheet, sheet_usuarios
    except Exception as e:
        logger.error(f"Erro ao conectar com Google Sheets: {e}")
        st.error("⚠ Erro ao conectar com Google Sheets. Verifique as credenciais e nome da planilha.")
        st.stop()

sheet, sheet_usuarios = conectar_google()

# ==============================
# FUNÇÕES DE DADOS OTIMIZADAS
# ==============================

@st.cache_data(ttl=CACHE_TTL_LONGO)
@retry_on_failure(max_retries=3)
def carregar_usuarios():
    """Carrega usuários com cache longo (raramente mudam)"""
    logger.info("Carregando usuários do Google Sheets")
    dados = sheet_usuarios.get_all_values()
    
    usuarios_dict = {}
    if len(dados) > 1:
        for linha in dados[1:]:
            if len(linha) >= 3:
                usuarios_dict[linha[0]] = {
                    "senha": linha[1],
                    "perfil": linha[2]
                }
    
    return usuarios_dict

@st.cache_data(ttl=CACHE_TTL_CURTO)
@retry_on_failure(max_retries=3)
def carregar_base():
    """Carrega base de dados com cache otimizado"""
    logger.info("Carregando base de dados do Google Sheets")
    dados = sheet.get_all_values()
    
    # Se não tem dados ou só tem cabeçalho, retorna DataFrame vazio com colunas corretas
    if len(dados) == 0:
        return pd.DataFrame(columns=["CCB", "Valor", "Parceiro", "Data da Análise", "Status Bankerize", "Status Analista", "Analista", "Anotações"])
    
    if len(dados) == 1:
        # Só tem cabeçalho, retorna DataFrame vazio mas com as colunas do cabeçalho
        return pd.DataFrame(columns=dados[0])
    
    header = dados[0]
    registros = dados[1:]
    df = pd.DataFrame(registros, columns=header)
    
    return df

def buscar_ccb_local(ccb, df=None):
    """Busca CCB no DataFrame local (sem acessar Google Sheets)"""
    if df is None:
        df = carregar_base()
    
    if df.empty:
        return None
    
    resultado = df[df["CCB"] == str(ccb)]
    
    if resultado.empty:
        return None
    
    return resultado.iloc[0]

@retry_on_failure(max_retries=3)
def assumir_ccb(ccb, valor, parceiro, analista, status_bankerize):
    """Assume CCB com validações e tratamento de erros"""
    if not ccb:
        return "Informe a CCB."
    
    # Busca no cache primeiro
    df = carregar_base()
    registro = df[df["CCB"] == str(ccb)]
    
    if not registro.empty:
        status = registro.iloc[0]["Status Analista"]
        
        if status in ["Análise Aprovada", "Análise Reprovada"]:
            return "⚠ Esta CCB já foi finalizada."
        
        if status in ["Em Análise", "Análise Pendente"]:
            st.session_state["ccb_ativa"] = ccb
            return "CONTINUAR"
    
    # Adiciona nova linha
    nova_linha = [
        ccb,
        valor,
        parceiro,
        datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%d/%m/%Y %H:%M:%S"),
        status_bankerize,
        "Em Análise",
        analista,
        ""
    ]
    
    try:
        sheet.append_row(nova_linha)
        # Limpa apenas o cache da base, não todos os caches
        carregar_base.clear()
        st.session_state["ccb_ativa"] = ccb
        logger.info(f"CCB {ccb} assumida por {analista}")
        return "OK"
    except Exception as e:
        logger.error(f"Erro ao assumir CCB: {e}")
        return f"Erro ao salvar: {str(e)}"

@retry_on_failure(max_retries=3)
def finalizar_ccb(ccb, resultado, anotacoes, status_bankerize):
    """Finaliza CCB usando batch update para melhor performance"""
    df = carregar_base()
    
    for idx, linha in df.iterrows():
        if str(linha["CCB"]) == str(ccb):
            linha_real = idx + 2
            
            try:
                # Usa batch update para reduzir requisições
                updates = [
                    {
                        'range': f'E{linha_real}',
                        'values': [[status_bankerize]]
                    },
                    {
                        'range': f'F{linha_real}',
                        'values': [[resultado]]
                    },
                    {
                        'range': f'H{linha_real}',
                        'values': [[anotacoes]]
                    }
                ]
                
                sheet.batch_update(updates)
                carregar_base.clear()
                logger.info(f"CCB {ccb} finalizada com status {resultado}")
                return "Finalizado"
            except Exception as e:
                logger.error(f"Erro ao finalizar CCB: {e}")
                return f"Erro ao finalizar: {str(e)}"
    
    return "CCB não encontrada."

@retry_on_failure(max_retries=3)
def adicionar_usuario(usuario, senha, perfil):
    """Adiciona usuário com senha criptografada"""
    try:
        senha_hash = hash_senha(senha)
        sheet_usuarios.append_row([usuario, senha_hash, perfil])
        carregar_usuarios.clear()
        logger.info(f"Usuário {usuario} adicionado com perfil {perfil}")
        return True
    except Exception as e:
        logger.error(f"Erro ao adicionar usuário: {e}")
        return False

@retry_on_failure(max_retries=3)
def excluir_usuario(usuario):
    """Exclui usuário da planilha"""
    try:
        dados = sheet_usuarios.get_all_values()
        for idx, linha in enumerate(dados[1:], start=2):
            if linha[0] == usuario:
                sheet_usuarios.delete_rows(idx)
                carregar_usuarios.clear()
                logger.info(f"Usuário {usuario} excluído")
                return True
        return False
    except Exception as e:
        logger.error(f"Erro ao excluir usuário: {e}")
        return False

@retry_on_failure(max_retries=3)
def atualizar_ccb(ccb, valor, parceiro, analista, status_bankerize):
    """Atualiza todos os dados de uma CCB existente na planilha"""
    df = carregar_base()
    for idx, linha in df.iterrows():
        if str(linha["CCB"]) == str(ccb):
            linha_real = idx + 2
            try:
                updates = [
                    {'range': f'B{linha_real}', 'values': [[valor]]},
                    {'range': f'C{linha_real}', 'values': [[parceiro]]},
                    {'range': f'D{linha_real}', 'values': [[datetime.now(ZoneInfo("America/Sao_Paulo")).strftime("%d/%m/%Y %H:%M:%S")]]},
                    {'range': f'E{linha_real}', 'values': [[status_bankerize]]},
                    {'range': f'F{linha_real}', 'values': [["Em Análise"]]},
                    {'range': f'G{linha_real}', 'values': [[analista]]},
                    {'range': f'H{linha_real}', 'values': [[""]]},
                ]
                sheet.batch_update(updates)
                carregar_base.clear()
                st.session_state["ccb_ativa"] = ccb
                logger.info(f"CCB {ccb} atualizada por {analista}")
                return "ATUALIZADO"
            except Exception as e:
                logger.error(f"Erro ao atualizar CCB: {e}")
                return f"Erro ao atualizar: {str(e)}"
    return "CCB não encontrada."

def alterar_senha(usuario, senha_atual, nova_senha):
    """Altera a senha de um usuário"""
    try:
        dados = sheet_usuarios.get_all_values()
        for idx, linha in enumerate(dados[1:], start=2):
            if linha[0] == usuario:
                if not verificar_senha(senha_atual, linha[1]):
                    return False, "Senha atual incorreta."
                novo_hash = hash_senha(nova_senha)
                sheet_usuarios.update_cell(idx, 2, novo_hash)
                carregar_usuarios.clear()
                logger.info(f"Senha alterada para usuário {usuario}")
                return True, "Senha alterada com sucesso."
        return False, "Usuário não encontrado."
    except Exception as e:
        logger.error(f"Erro ao alterar senha: {e}")
        return False, f"Erro: {str(e)}"

# ==============================
# LOGIN
# ==============================

def login():
    st.markdown("""
    <style>
    [data-testid="stSidebar"],
    [data-testid="collapsedControl"] { display: none !important; }
    .main .block-container {
        max-width: 460px !important;
        padding: 10vh 1rem 4rem !important;
        margin: 0 auto !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Logo: usa base64 para funcionar no Render sem depender de path relativo
    import base64
    logo_html = ""
    try:
        with open("Logo Principal.png", "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()
        logo_html = f'<img src="data:image/png;base64,{logo_b64}" style="max-width:200px;height:auto;border-radius:10px;margin:0 auto 1.25rem;display:block;" />'
    except Exception:
        logo_html = """
        <div style="display:inline-flex;flex-direction:column;align-items:center;
            background:linear-gradient(135deg,#22c55e,#16a34a);border-radius:14px;
            padding:1rem 2rem;margin:0 auto 1.25rem;width:fit-content;
            box-shadow:0 8px 24px rgba(34,197,94,0.3);">
            <span style="font-size:1.75rem;font-weight:800;color:#fff;letter-spacing:0.06em;line-height:1;">CONSIGA</span>
            <span style="font-size:0.75rem;font-weight:600;color:#fbbf24;letter-spacing:0.12em;text-transform:uppercase;margin-top:4px;">Empréstimos</span>
        </div>"""

    st.markdown(f"""
    <div style="background:var(--bg-primary);border-radius:16px;padding:2rem 2rem 1.5rem;
        box-shadow:0 8px 32px rgba(0,0,0,0.12);border:1px solid var(--border-color);text-align:center;margin-bottom:1.25rem;">
        {logo_html}
        <p style="font-size:1.25rem;font-weight:700;margin:0 0 3px;">Bem-vindo</p>
        <p style="font-size:0.8125rem;font-weight:400;margin:0;opacity:0.6;">Sistema de Controle de Análise de Crédito</p>
    </div>
    """, unsafe_allow_html=True)

    user     = st.text_input("Usuário", placeholder="Digite seu usuário",  key="login_user")
    password = st.text_input("Senha",   placeholder="Digite sua senha",    key="login_pass", type="password")

    if st.button("Entrar", use_container_width=True, key="login_btn"):
        try:
            usuarios = carregar_usuarios()
            if user in usuarios and verificar_senha(password, usuarios[user]["senha"]):
                st.session_state["user"]   = user
                st.session_state["perfil"] = usuarios[user]["perfil"]
                logger.info(f"Login bem-sucedido: {user}")
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos")
                logger.warning(f"Tentativa de login falhou: {user}")
        except Exception as e:
            st.error("Erro ao fazer login. Tente novamente.")
            logger.error(f"Erro no login: {e}")


if "user" not in st.session_state:
    login()
    st.stop()

# Após login: restaura layout normal — sidebar sempre visível, botão de colapsar NUNCA aparece
st.markdown("""
<style>
[data-testid="stSidebar"] { display: flex !important; }

/* Garante que o botão de colapsar nunca apareça em nenhuma situação */
[data-testid="collapsedControl"],
[data-testid="stSidebarCollapseButton"],
button[data-testid="baseButton-header"],
section[data-testid="stSidebar"] > div > button {
    display: none !important;
    visibility: hidden !important;
    pointer-events: none !important;
    width: 0 !important;
    height: 0 !important;
    position: absolute !important;
    left: -9999px !important;
}

.main .block-container {
    max-width: 100% !important;
    padding: 2rem 2rem 4rem !important;
}
</style>
""", unsafe_allow_html=True)

analista = st.session_state["user"]

# ==============================
# MENU LATERAL
# ==============================

opcoes_menu = ["Operação"]

if st.session_state["perfil"] == "Supervisor":
    opcoes_menu.append("Acompanhamento")
    opcoes_menu.append("Administração")

# Logo na sidebar com base64 (funciona no Render)
import base64 as _b64
try:
    with open("Logo Principal.png", "rb") as _f:
        _logo_b64 = _b64.b64encode(_f.read()).decode()
    st.sidebar.markdown(f"""
    <div style="padding:1.25rem 1rem 0.75rem;text-align:center;">
        <img src="data:image/png;base64,{_logo_b64}"
             style="max-width:150px;width:100%;height:auto;border-radius:6px;display:block;margin:0 auto;" />
    </div>
    """, unsafe_allow_html=True)
except Exception:
    st.sidebar.markdown("""
    <div style="padding:1.25rem 1rem 0.75rem;text-align:center;">
        <div style="display:inline-flex;flex-direction:column;align-items:center;
            background:rgba(34,197,94,0.15);border:1px solid rgba(34,197,94,0.3);
            border-radius:10px;padding:0.6rem 1.2rem;">
            <span style="font-size:1.1rem;font-weight:800;color:#fff;letter-spacing:0.06em;">CONSIGA</span>
            <span style="font-size:0.6rem;font-weight:600;color:#fb923c;letter-spacing:0.1em;text-transform:uppercase;margin-top:2px;">Empréstimos</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

st.sidebar.markdown("---")
menu = st.sidebar.selectbox("Navegação", opcoes_menu, label_visibility="collapsed")
st.sidebar.markdown("---")
st.sidebar.markdown(f"**Usuário:** {analista}")
st.sidebar.markdown(f"**Perfil:** {st.session_state['perfil']}")
st.sidebar.markdown("")

# Botão de atualização manual
if st.sidebar.button("Atualizar Dados", use_container_width=True):
    carregar_base.clear()
    st.rerun()

st.sidebar.markdown("")
if st.sidebar.button("Sair", use_container_width=True, type="primary"):
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

# Troca de senha própria (disponível para todos os perfis via expander na sidebar)
with st.sidebar.expander("Alterar minha senha"):
    s_atual  = st.text_input("Senha atual",    type="password", key="s_atual")
    s_nova   = st.text_input("Nova senha",     type="password", key="s_nova")
    s_nova2  = st.text_input("Confirmar nova", type="password", key="s_nova2")
    if st.button("Salvar senha", key="btn_salvar_senha", use_container_width=True):
        if not s_atual or not s_nova:
            st.error("Preencha todos os campos.")
        elif s_nova != s_nova2:
            st.error("As senhas não coincidem.")
        elif len(s_nova) < 4:
            st.error("A senha deve ter ao menos 4 caracteres.")
        else:
            ok, msg = alterar_senha(analista, s_atual, s_nova)
            if ok:
                st.success(msg)
            else:
                st.error(msg)

# ==============================
# OPERAÇÃO
# ==============================

if menu == "Operação":
    
    st.title("Mesa de Análise CCB")
    
    ccb_input = st.text_input("Número da CCB")
    valor = st.text_input("Valor Líquido")
    parceiro = st.text_input("Parceiro")
    
    status_bankerize = st.selectbox(
        "Status Bankerize",
        [
            "Aguardando Análise da Assinatura",
            "Aguardando Análise de Risco",
            "Aguardando Análise Manual da Assinatura",
            "Assinatura Reprovada",
            "Pendente"
        ],
        key="status_bankerize_select"
    )

    # ==============================
    # BUSCA EM TEMPO REAL
    # ==============================
    info = None
    if ccb_input:
        df_cache = carregar_base()
        info = buscar_ccb_local(ccb_input, df_cache)
        if info is not None:
            status_existente = info["Status Analista"]
            cor = {"Análise Aprovada": "#22c55e", "Análise Reprovada": "#ef4444",
                   "Em Análise": "#3b82f6", "Análise Pendente": "#f97316"}.get(status_existente, "#64748b")
            st.markdown(f"""
            <div style="border-left:4px solid {cor};background:rgba(0,0,0,0.04);
                border-radius:8px;padding:12px 16px;margin:8px 0;">
                <strong>CCB já cadastrada</strong><br>
                Analista: {info['Analista']} &nbsp;|&nbsp;
                Status: <span style="color:{cor};font-weight:600;">{status_existente}</span><br>
                Data: {info['Data da Análise']}
            </div>
            """, unsafe_allow_html=True)

    # ==============================
    # BOTÃO ASSUMIR + LÓGICA DE ATUALIZAÇÃO
    # ==============================
    if st.button("Assumir Análise"):
        if not ccb_input:
            st.error("Informe o número da CCB.")
        elif info is not None:
            # CCB já existe — pergunta se quer atualizar
            st.session_state["ccb_confirmacao_pendente"] = {
                "ccb": ccb_input, "valor": valor, "parceiro": parceiro,
                "status_bankerize": status_bankerize
            }
            st.rerun()

        else:
            with st.spinner("Processando..."):
                resposta = assumir_ccb(ccb_input, valor, parceiro, analista, status_bankerize)
                if resposta == "OK":
                    st.success("CCB criada e assumida com sucesso")
                    st.rerun()
                elif resposta == "CONTINUAR":
                    st.success("Retomando análise desta CCB")
                else:
                    st.error(resposta)

    # Modal de confirmação de atualização
    if "ccb_confirmacao_pendente" in st.session_state:
        dados_pendentes = st.session_state["ccb_confirmacao_pendente"]
        st.warning(f"A CCB **{dados_pendentes['ccb']}** já existe. Deseja atualizá-la com os novos dados?")
        col_sim, col_nao = st.columns(2)
        with col_sim:
            if st.button("Sim, atualizar", use_container_width=True):
                with st.spinner("Atualizando..."):
                    resp = atualizar_ccb(
                        dados_pendentes["ccb"], dados_pendentes["valor"],
                        dados_pendentes["parceiro"], analista,
                        dados_pendentes["status_bankerize"]
                    )
                    del st.session_state["ccb_confirmacao_pendente"]
                    if resp == "ATUALIZADO":
                        st.success("CCB atualizada com sucesso")
                        st.rerun()
                    else:
                        st.error(resp)
        with col_nao:
            if st.button("Não, continuar sem atualizar", use_container_width=True):
                ccb_retomar = dados_pendentes["ccb"]
                del st.session_state["ccb_confirmacao_pendente"]
                st.session_state["ccb_ativa"] = ccb_retomar
                st.rerun()

    if "ccb_ativa" in st.session_state:
        st.divider()
        st.subheader(f"Finalizando CCB {st.session_state['ccb_ativa']}")
        
        resultado = st.radio(
            "Resultado",
            ["Análise Pendente", "Análise Aprovada", "Análise Reprovada"]
        )
        anotacoes = st.text_area("Anotações")
        
        if st.button("Finalizar Análise"):
            if resultado == "Análise Pendente" and not anotacoes:
                st.error("Para Análise Pendente é obrigatório preencher Anotações.")
            else:
                with st.spinner("Finalizando..."):
                    resultado_final = finalizar_ccb(
                        st.session_state["ccb_ativa"],
                        resultado,
                        anotacoes,
                        status_bankerize
                    )
                    if resultado_final == "Finalizado":
                        st.success("Análise finalizada com sucesso")
                        del st.session_state["ccb_ativa"]
                        st.rerun()
                    else:
                        st.error(resultado_final)

    # ==============================
    # PAINEL GERAL COM FILTROS
    # ==============================
    st.divider()
    st.subheader("Painel Geral")

    df = carregar_base().copy()

    if not df.empty:
        df["Data da Análise"] = pd.to_datetime(df["Data da Análise"], dayfirst=True, errors="coerce")
        df = df.dropna(subset=["Data da Análise"])
        df = df.sort_values(by="Data da Análise", ascending=False)

        # Filtros
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            analistas_disponiveis = ["Todos"] + sorted(df["Analista"].dropna().unique().tolist())
            filtro_analista = st.selectbox("Filtrar por Analista", analistas_disponiveis, key="filtro_analista_painel")
        with col_f2:
            status_disponiveis = ["Todos"] + sorted(df["Status Analista"].dropna().unique().tolist())
            filtro_status = st.selectbox("Filtrar por Status", status_disponiveis, key="filtro_status_painel")

        df_filtrado = df.copy()
        if filtro_analista != "Todos":
            df_filtrado = df_filtrado[df_filtrado["Analista"] == filtro_analista]
        if filtro_status != "Todos":
            df_filtrado = df_filtrado[df_filtrado["Status Analista"] == filtro_status]

        df_filtrado["Data da Análise"] = df_filtrado["Data da Análise"].dt.strftime("%d/%m/%Y %H:%M:%S")
        st.caption(f"{len(df_filtrado)} registro(s) encontrado(s)")
        st.dataframe(df_filtrado, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhum registro encontrado.")

# ==============================
# ACOMPANHAMENTO (apenas Supervisor)
# ==============================

if menu == "Acompanhamento":

    if st.session_state["perfil"] != "Supervisor":
        st.warning("Acesso restrito a Supervisores.")
        st.stop()

    st.title("Acompanhamento")

    df_raw = carregar_base().copy()

    if df_raw.empty:
        st.info("Nenhum registro encontrado.")
    else:
        df_raw["Data da Análise"] = pd.to_datetime(df_raw["Data da Análise"], dayfirst=True, errors="coerce")
        df_raw = df_raw.dropna(subset=["Data da Análise"])

        # ==============================
        # FILTRO LIVRE POR DATA
        # ==============================
        hoje = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
        inicio_mes = hoje.replace(day=1)

        col_f1, col_f2 = st.columns(2)
        with col_f1:
            data_inicio = st.date_input("Data de início", value=inicio_mes, format="DD/MM/YYYY")
        with col_f2:
            data_fim = st.date_input("Data de fim", value=hoje, format="DD/MM/YYYY")

        df_raw["Data_date"] = df_raw["Data da Análise"].dt.date
        df = df_raw[(df_raw["Data_date"] >= data_inicio) & (df_raw["Data_date"] <= data_fim)].copy()

        st.caption(f"Exibindo registros de {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')} — {len(df)} registro(s)")

        if df.empty:
            st.info("Nenhuma proposta encontrada no período selecionado.")
        else:
            # ==============================
            # MÉTRICAS DO PERÍODO
            # ==============================
            pendentes  = df[df["Status Analista"] == "Análise Pendente"].shape[0]
            aprovadas  = df[df["Status Analista"] == "Análise Aprovada"].shape[0]
            reprovadas = df[df["Status Analista"] == "Análise Reprovada"].shape[0]
            em_analise = df[df["Status Analista"] == "Em Análise"].shape[0]
            total      = len(df)

            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Total", total)
            col2.metric("Em Análise", em_analise)
            col3.metric("Pendentes", pendentes)
            col4.metric("Aprovadas", aprovadas)
            col5.metric("Reprovadas", reprovadas)

            st.divider()

            # ==============================
            # GRÁFICO DO PERÍODO
            # ==============================
            fig, ax = plt.subplots(figsize=(9, 4))
            labels = ["Em Análise", "Pendentes", "Aprovadas", "Reprovadas"]
            valores = [em_analise, pendentes, aprovadas, reprovadas]
            cores   = ["#3b82f6", "#f97316", "#22c55e", "#ef4444"]
            barras  = ax.bar(labels, valores, color=cores, width=0.45, edgecolor="none")
            for b in barras:
                h = b.get_height()
                if h > 0:
                    ax.text(b.get_x() + b.get_width()/2, h + 0.05,
                            str(int(h)), ha='center', va='bottom', fontsize=12, fontweight='bold')
            ax.set_ylabel("Quantidade", fontsize=10)
            ax.set_title(f"Propostas — {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
                         fontsize=12, fontweight='bold', pad=12)
            ax.spines[['top','right']].set_visible(False)
            ax.set_axisbelow(True)
            ax.yaxis.grid(True, linestyle='--', alpha=0.4)
            plt.tight_layout()
            st.pyplot(fig)
            plt.close(fig)

            st.divider()

            # ==============================
            # DASHBOARD POR ANALISTA
            # ==============================
            st.subheader("Por Analista")

            resumo = df.groupby("Analista").agg(
                Total=("Status Analista", "count"),
                Em_Analise=("Status Analista", lambda x: (x == "Em Análise").sum()),
                Pendentes=("Status Analista", lambda x: (x == "Análise Pendente").sum()),
                Aprovadas=("Status Analista", lambda x: (x == "Análise Aprovada").sum()),
                Reprovadas=("Status Analista", lambda x: (x == "Análise Reprovada").sum())
            ).reset_index().sort_values("Total", ascending=False)

            st.dataframe(resumo, use_container_width=True, hide_index=True)

            st.divider()

            # ==============================
            # RANKING DE PRODUTIVIDADE
            # ==============================
            st.subheader("Ranking de Produtividade")

            ranking = df[df["Status Analista"].isin(["Análise Aprovada", "Análise Reprovada"])]\
                .groupby("Analista").size().reset_index(name="Finalizadas")\
                .sort_values("Finalizadas", ascending=False).reset_index(drop=True)

            if not ranking.empty:
                medalhas = ["🥇", "🥈", "🥉"]
                for i, row in ranking.iterrows():
                    medalha = medalhas[i] if i < 3 else f"{i+1}."
                    pct = int(row["Finalizadas"] / ranking["Finalizadas"].sum() * 100)
                    st.markdown(f"""
                    <div style="display:flex;align-items:center;gap:12px;
                        padding:10px 16px;border-radius:10px;margin-bottom:6px;
                        background:{'rgba(34,197,94,0.12)' if i == 0 else 'rgba(0,0,0,0.03)'};
                        border:1px solid {'rgba(34,197,94,0.3)' if i == 0 else 'rgba(0,0,0,0.06)'};">
                        <span style="font-size:1.4rem;">{medalha}</span>
                        <span style="font-weight:700;flex:1;">{row['Analista']}</span>
                        <span style="font-weight:600;color:#16a34a;">{int(row['Finalizadas'])} finalizadas</span>
                        <span style="font-size:0.8rem;color:#64748b;">{pct}%</span>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("Nenhuma análise finalizada no período.")

            st.divider()

            # ==============================
            # RELATÓRIO DETALHADO + EXPORTAR EXCEL
            # ==============================
            st.subheader("Relatório Detalhado")

            df_export = df.copy()
            df_export["Data da Análise"] = df_export["Data da Análise"].dt.strftime("%d/%m/%Y %H:%M:%S")
            df_export = df_export.drop(columns=["Data_date"], errors="ignore")
            df_export = df_export.sort_values("Data da Análise", ascending=False)

            col_tab, col_btn = st.columns([5, 1])
            with col_tab:
                st.dataframe(df_export, use_container_width=True, hide_index=True)
            with col_btn:
                # Exportar Excel formatado
                import io
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Relatório"

                # Cabeçalho verde
                header_fill = PatternFill("solid", fgColor="16A34A")
                header_font = Font(bold=True, color="FFFFFF", size=10)
                thin = Side(style="thin", color="D1D5DB")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for col_idx, col_name in enumerate(df_export.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                # Dados
                alt_fill = PatternFill("solid", fgColor="F1F5F9")
                for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(vertical="center")
                        if row_idx % 2 == 0:
                            cell.fill = alt_fill

                # Ajusta largura das colunas
                for col_idx, col_name in enumerate(df_export.columns, 1):
                    max_len = max(len(str(col_name)), df_export[col_name].astype(str).str.len().max())
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

                ws.row_dimensions[1].height = 20

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                st.download_button(
                    label="Exportar Excel",
                    data=buf,
                    file_name=f"relatorio_{data_inicio.strftime('%d%m%Y')}_{data_fim.strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

# ==============================
# ADMINISTRAÇÃO
# ==============================

if menu == "Administração":
    
    if st.session_state["perfil"] != "Supervisor":
        st.warning("Acesso restrito a Supervisores.")
        st.stop()
    
    st.title("Administração de Usuários")
    
    usuarios = carregar_usuarios()
    
    # LISTAR USUÁRIOS
    lista = []
    for nome, dados in usuarios.items():
        lista.append({
            "Usuário": nome,
            "Perfil": dados["perfil"]
        })
    
    df_users = pd.DataFrame(lista)
    st.dataframe(df_users, use_container_width=True, hide_index=True)
    
    st.divider()
    
    # ADICIONAR USUÁRIO
    st.subheader("Adicionar Novo Usuário")
    
    novo_usuario = st.text_input("Nome do Usuário")
    nova_senha = st.text_input("Senha", type="password")
    nova_senha_confirm = st.text_input("Confirmar Senha", type="password")
    novo_perfil = st.selectbox("Perfil", ["Operador", "Supervisor"])
    
    if st.button("Cadastrar Usuário"):
        
        if not novo_usuario or not nova_senha:
            st.error("Preencha todos os campos.")
        elif nova_senha != nova_senha_confirm:
            st.error("As senhas não coincidem.")
        elif novo_usuario in usuarios:
            st.error("Usuário já existe.")
        else:
            with st.spinner("Cadastrando usuário..."):
                if adicionar_usuario(novo_usuario, nova_senha, novo_perfil):
                    st.success("Usuário cadastrado com sucesso")
                    st.rerun()
                else:
                    st.error("Erro ao cadastrar usuário.")
    
    st.divider()
    
    # EXCLUIR USUÁRIO
    st.subheader("Excluir Usuário")
    
    usuario_excluir = st.selectbox(
        "Selecionar Usuário para Excluir",
        list(usuarios.keys())
    )
    
    if st.button("Excluir Usuário", type="primary"):
        if usuario_excluir == analista:
            st.error("Você não pode excluir seu próprio usuário.")
        else:
            with st.spinner("Excluindo usuário..."):
                if excluir_usuario(usuario_excluir):
                    st.success("Usuário excluído com sucesso")
                    st.rerun()
                else:
                    st.error("Erro ao excluir usuário.")

    st.divider()

    # TROCA DE SENHA (qualquer usuário pode trocar a própria senha)
    st.subheader("Alterar Senha de Usuário")

    usuario_senha = st.selectbox("Usuário", list(usuarios.keys()), key="sel_usuario_senha")
    senha_nova    = st.text_input("Nova Senha", type="password", key="nova_senha_adm")
    senha_nova2   = st.text_input("Confirmar Nova Senha", type="password", key="nova_senha_adm2")

    if st.button("Alterar Senha", key="btn_alterar_senha_adm"):
        if not senha_nova:
            st.error("Informe a nova senha.")
        elif senha_nova != senha_nova2:
            st.error("As senhas não coincidem.")
        else:
            with st.spinner("Alterando senha..."):
                novo_hash = hash_senha(senha_nova)
                try:
                    dados_u = sheet_usuarios.get_all_values()
                    for idx_u, linha_u in enumerate(dados_u[1:], start=2):
                        if linha_u[0] == usuario_senha:
                            sheet_usuarios.update_cell(idx_u, 2, novo_hash)
                            carregar_usuarios.clear()
                            st.success(f"Senha de {usuario_senha} alterada com sucesso.")
                            break
                except Exception as e:
                    st.error(f"Erro ao alterar senha: {e}")
